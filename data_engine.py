"""
data_engine.py  v2
------------------
Changes from v1:
- P1 start = next BD after last BD of November (prev year)
- Rogers rollover dates in bounds dict
- add_metrics(): close_price, daily_range, TR, ATR, px_chg, oi_chg, vol_chg, week_in_period
- directional_bias() returns (df, stats_dict) — no df.attrs
- Fly computed cleanly with proper column names
- bd_to_date_map(), current_bd_info()
- streak_analysis(), return_dist(), autocorr_stats(), cross_corr(), week_stats()
"""

import pandas as pd
import numpy as np
from datetime import date, timedelta
from scipy import stats
import openpyxl
from datetime import date as date_type

DEFAULT_PATHS = {
    "outright":  "CC_H_yearwise_workbook.xlsx",
    "hk_spread": "CCH_spread_yearwise_workbook.xlsx",
    "kn_spread":  "CCK_spread_yearwise_workbook.xlsx",
}

# ── Business day utilities ────────────────────────────────────────────────────
def is_bd(d): return d.weekday() < 5

def nth_bd(year, month, n):
    d = date(year, month, 1); c = 0
    while True:
        if is_bd(d): c += 1
        if c == n: return d
        d += timedelta(1)

def last_bd(year, month):
    nxt = date(year+1,1,1) if month==12 else date(year,month+1,1)
    d = nxt - timedelta(1)
    while not is_bd(d): d -= timedelta(1)
    return d

def prev_bd(d):
    d -= timedelta(1)
    while not is_bd(d): d -= timedelta(1)
    return d

def next_bd(d):
    d += timedelta(1)
    while not is_bd(d): d += timedelta(1)
    return d

def count_bd(start, end):
    if end < start: return 0
    return sum(1 for i in range((end-start).days+1) if is_bd(start+timedelta(i)))

def cc_mar_ltd(year):
    lbd = last_bd(year, 3); d = lbd
    for _ in range(10): d = prev_bd(d)
    return prev_bd(d)

# ── Period boundaries ─────────────────────────────────────────────────────────
def get_boundaries(year):
    """
    P1: next BD after last BD of November (prev year) → day before Goldman roll
    P2: 5th BD of Feb → 9th BD of Feb  (Goldman roll)
    P3: day after Goldman end → CC March LTD
    Rogers IN roll: penultimate BD Nov → 1st BD Dec (rolling INTO Dec, front becomes H)
    Rogers OUT roll: penultimate BD Feb → 1st BD Mar (rolling OUT of H → K)
    """
    nov_last   = last_bd(year-1, 11)
    p1_start   = next_bd(nov_last)
    gs         = nth_bd(year, 2, 5)
    ge         = nth_bd(year, 2, 9)
    p1_end     = prev_bd(gs)
    p3_start   = next_bd(ge)
    mar_ltd    = cc_mar_ltd(year)

    # Rogers IN
    r_in_s  = prev_bd(nov_last)          # penultimate BD of Nov
    r_in_e  = nth_bd(year-1, 12, 1)     # 1st BD of Dec

    # Rogers OUT
    feb_last = last_bd(year, 2)
    r_out_s  = prev_bd(feb_last)
    r_out_e  = nth_bd(year, 3, 1)

    return dict(
        year=year,
        nov_last_bd=nov_last,
        p1_start=p1_start, p1_end=p1_end,
        p2_start=gs,       p2_end=ge,
        p3_start=p3_start, p3_end=mar_ltd,
        mar_ltd=mar_ltd,
        p1_bd=count_bd(p1_start, p1_end),
        p3_bd=count_bd(p3_start, mar_ltd),
        rogers_in_start=r_in_s,  rogers_in_end=r_in_e,
        rogers_out_start=r_out_s, rogers_out_end=r_out_e,
    )

# ── Workbook loader ───────────────────────────────────────────────────────────
def load_wb(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    out = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2: continue
        df = pd.DataFrame(rows[1:], columns=rows[0])
        df['Timestamp'] = pd.to_datetime(df['Timestamp'], errors='coerce')
        df = df.dropna(subset=['Timestamp'])
        for col in ['Open','High','Low','Last','Settlement Price','Open Interest','Volume','VWAP']:
            if col in df.columns: df[col] = pd.to_numeric(df[col], errors='coerce')
        df = df.sort_values('Timestamp').reset_index(drop=True)
        out[sheet] = df
    return out

# ── Period tagger ─────────────────────────────────────────────────────────────
def tag_periods(df, bounds):
    df = df.copy()
    df['date'] = df['Timestamp'].dt.date

    def _p(d):
        if bounds['p1_start'] <= d <= bounds['p1_end']: return 'P1'
        if bounds['p2_start'] <= d <= bounds['p2_end']: return 'P2'
        if bounds['p3_start'] <= d <= bounds['p3_end']: return 'P3'
        return None

    df['period'] = df['date'].apply(_p)
    df = df[df['period'].notna()].copy()

    for p in ['P1','P2','P3']:
        idx = df[df['period']==p].sort_values('date').index
        df.loc[idx, 'bd'] = range(1, len(idx)+1)

    df['bd'] = df['bd'].astype('Int64')
    df['bd_to_ltd'] = df['date'].apply(lambda d: count_bd(d, bounds['mar_ltd']))
    df['week_in_period'] = ((df['bd'].astype(float) - 1) // 5 + 1).astype('Int64')
    return df

# ── Metrics ───────────────────────────────────────────────────────────────────
def add_metrics(df):
    """Add close_price, daily_range, TR, ATR(14), px_chg, oi_chg, vol_chg."""
    df = df.copy().sort_values(['period','date'])

    if 'Last' in df.columns and 'Settlement Price' in df.columns:
        df['close_price'] = df['Last'].fillna(df['Settlement Price'])
    elif 'Settlement Price' in df.columns:
        df['close_price'] = df['Settlement Price']
    elif 'Last' in df.columns:
        df['close_price'] = df['Last']
    else:
        df['close_price'] = np.nan

    if 'High' in df.columns and 'Low' in df.columns:
        df['daily_range'] = (df['High'] - df['Low']).where(
            df['High'].notna() & df['Low'].notna())
        prev_c = df['close_price'].shift(1)
        hl = df['daily_range'].fillna(0)
        hc = (df['High'] - prev_c).abs()
        lc = (df['Low'] - prev_c).abs()
        df['TR'] = pd.concat([hl, hc, lc], axis=1).max(axis=1)
        df['ATR14'] = df['TR'].rolling(14, min_periods=3).mean()
    else:
        df['daily_range'] = np.nan
        df['TR'] = np.nan
        df['ATR14'] = np.nan

    df['px_chg']  = df.groupby('period')['close_price'].diff()
    df['px_pct']  = df.groupby('period')['close_price'].pct_change() * 100
    if 'Open Interest' in df.columns:
        df['oi_chg'] = df.groupby('period')['Open Interest'].diff()
    if 'Volume' in df.columns:
        df['vol_chg'] = df.groupby('period')['Volume'].diff()
    return df

# ── Fly computation ───────────────────────────────────────────────────────────
def compute_fly(hk_df, kn_df):
    if len(hk_df) == 0 or len(kn_df) == 0: return pd.DataFrame()

    hk = hk_df[hk_df['close_price'].notna()].copy()
    kn = kn_df[kn_df['close_price'].notna()].copy()

    cols_keep = ['date','period','bd','week_in_period','bd_to_ltd']

    m = pd.merge(
        hk[cols_keep + ['close_price','Volume','Open','High','Low']].rename(
            columns={'close_price':'hk','Volume':'vol_hk','Open':'o_hk',
                     'High':'h_hk','Low':'l_hk'}),
        kn[cols_keep + ['close_price','Volume']].rename(
            columns={'close_price':'kn','Volume':'vol_kn'}),
        on=cols_keep, how='inner'
    )
    m['close_price'] = m['hk'] - m['kn']   # HKN fly = HK_spread - KN_spread
    m['Volume']      = m[['vol_hk','vol_kn']].mean(axis=1)
    m['Open']        = m['o_hk'] - m['kn']   # approximate
    m['Timestamp']   = pd.to_datetime(m['date'])
    for c in ['High','Low','Last','Open Interest','daily_range','TR','ATR14']:
        m[c] = np.nan
    m['px_chg'] = m.groupby('period')['close_price'].diff()
    m['px_pct']  = m.groupby('period')['close_price'].pct_change() * 100
    m['oi_chg']  = np.nan
    m['vol_chg'] = m.groupby('period')['Volume'].diff()
    m['week_in_period'] = m['week_in_period'].astype('Int64')
    return m

# ── Main loader ───────────────────────────────────────────────────────────────
def load_all(paths=None):
    if paths is None: paths = DEFAULT_PATHS
    print("Loading workbooks…")
    raw_out = load_wb(paths['outright'])
    raw_hk  = load_wb(paths['hk_spread'])
    raw_kn  = load_wb(paths['kn_spread'])

    years = sorted([int(s) for s in raw_out if s.isdigit() and 2014 <= int(s) <= 2025])
    data  = {}

    for yr in years:
        s      = str(yr)
        bounds = get_boundaries(yr)

        def _prep(raw, sheet):
            if sheet not in raw: return pd.DataFrame()
            df = tag_periods(raw[sheet], bounds)
            return add_metrics(df)

        out = _prep(raw_out, s)
        hk  = _prep(raw_hk,  s)
        kn  = _prep(raw_kn,  s)
        fly = compute_fly(hk, kn) if len(hk) and len(kn) else pd.DataFrame()

        data[yr] = dict(bounds=bounds, out=out, hk=hk, kn=kn, fly=fly)

    print(f"Loaded {len(years)} years: {years[0]}–{years[-1]}")
    return data, years

# ── Analysis helpers ──────────────────────────────────────────────────────────

def directional_bias(data, years, series='out', col='close_price', period='P1'):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        p = sub[(sub['period']==period) & sub[col].notna()].sort_values('date')
        if len(p) < 2: continue
        f, l = p[col].iloc[0], p[col].iloc[-1]
        rows.append(dict(year=yr, first=f, last=l,
                         net_move=l-f, pct_move=(l-f)/abs(f)*100 if f!=0 else 0,
                         direction='UP' if l > f else 'DOWN'))
    df = pd.DataFrame(rows)
    stats_d = {}
    if len(df):
        up = (df['direction']=='UP').sum()
        stats_d = dict(
            hit_rate_up=up/len(df), hit_rate=max(up/len(df),1-up/len(df)),
            avg_move=df['net_move'].mean(), avg_pct=df['pct_move'].mean(),
            direction='UP' if up/len(df) >= 0.5 else 'DOWN', n=len(df),
            pval=stats.binomtest(int(up), len(df), 0.5).pvalue
        )
    return df, stats_d

def consistent_pattern_days(data, years, series='out', col='close_price',
                             period='P1', min_hit=0.65):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        p = sub[(sub['period']==period) & sub[col].notna()].sort_values('date').copy()
        if 'px_chg' not in p.columns: p['px_chg'] = p[col].diff()
        p['year'] = yr
        rows.append(p)
    if not rows: return []
    combined = pd.concat(rows, ignore_index=True)
    patterns = []
    for bd_pos, grp in combined.groupby('bd'):
        grp = grp[grp['px_chg'].notna()]
        if len(grp) < 6: continue
        up = (grp['px_chg'] > 0).sum(); n = len(grp)
        hr_up = up/n; hr_dn = 1-hr_up
        if hr_up >= min_hit:
            patterns.append(dict(bd=int(bd_pos), direction='UP', hit_rate=hr_up,
                                 avg_move=grp['px_chg'].mean(), n=n,
                                 years=sorted(grp['year'].tolist())))
        elif hr_dn >= min_hit:
            patterns.append(dict(bd=int(bd_pos), direction='DOWN', hit_rate=hr_dn,
                                 avg_move=grp['px_chg'].mean(), n=n,
                                 years=sorted(grp['year'].tolist())))
    return sorted(patterns, key=lambda x: -x['hit_rate'])

def seasonal_scorecard(data, years):
    rows = []
    for yr in [y for y in years if y in data]:
        b = data[yr]['bounds']
        row = {'year': yr, 'dec_ltd': b['nov_last_bd'], 'mar_ltd': b['mar_ltd'],
               'gs': b['p2_start'], 'ge': b['p2_end']}
        for series, label in [('out','out'),('hk','hk'),('fly','fly')]:
            sub = data[yr].get(series, pd.DataFrame())
            if len(sub) == 0:
                for p in ['P1','P2','P3']: row[f'{label}_{p}'] = None
                continue
            col = 'close_price'
            for p in ['P1','P2','P3']:
                pp = sub[(sub['period']==p) & sub[col].notna()].sort_values('date')
                if len(pp) < 2: row[f'{label}_{p}'] = None; continue
                net = pp[col].iloc[-1] - pp[col].iloc[0]
                row[f'{label}_{p}'] = '▲' if net > 0 else '▼'
        rows.append(row)
    return pd.DataFrame(rows)

def oi_cliff(data, years):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get('out', pd.DataFrame())
        if len(sub) == 0: continue
        p1 = sub[(sub['period']=='P1') & sub['Open Interest'].notna()]
        if len(p1) == 0: continue
        peak = p1['Open Interest'].max()
        all_oi = sub[sub['Open Interest'].notna()].sort_values('date')
        below = all_oi[all_oi['Open Interest'] < peak * 0.5]
        if len(below):
            cd = below['date'].iloc[0]
            rows.append({'year': yr, 'cliff_date': cd,
                         'bd_before_ltd': count_bd(cd, data[yr]['bounds']['mar_ltd']),
                         'peak_oi': peak, 'cliff_oi': below['Open Interest'].iloc[0]})
    return pd.DataFrame(rows)

def range_ratio(data, years, series='out'):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        for p in ['P1','P2','P3']:
            pp = sub[(sub['period']==p) & sub['daily_range'].notna()]
            if len(pp) == 0: continue
            rows.append(dict(year=yr, period=p, avg_range=pp['daily_range'].mean(), n=len(pp)))
    return pd.DataFrame(rows)

def vol_oi_patterns(data, years, series='out', period='P1'):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        p = sub[(sub['period']==period) & sub['close_price'].notna()].sort_values('date').copy()
        p['year'] = yr
        rows.append(p)
    if not rows: return pd.DataFrame()
    combined = pd.concat(rows, ignore_index=True)
    agg = combined.groupby('bd').agg(
        avg_px_chg=('px_chg','mean'), std_px_chg=('px_chg','std'),
        avg_vol=('Volume','mean'), avg_oi_chg=('oi_chg','mean'),
        n_years=('year','count'),
        up_days=('px_chg', lambda x: (x>0).sum()),
    ).reset_index()
    agg['hit_rate_up'] = agg['up_days'] / agg['n_years']
    return agg

def streak_analysis(data, years, series='out', period='P1'):
    """Count consecutive up/down streak frequencies."""
    all_streaks = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        p = sub[(sub['period']==period) & sub['px_chg'].notna()].sort_values('date')
        directions = (p['px_chg'] > 0).astype(int).tolist()
        streak, prev = 1, directions[0] if directions else None
        for d in directions[1:]:
            if d == prev: streak += 1
            else:
                all_streaks.append({'year': yr, 'direction': 'UP' if prev else 'DOWN', 'length': streak})
                streak = 1; prev = d
        if prev is not None:
            all_streaks.append({'year': yr, 'direction': 'UP' if prev else 'DOWN', 'length': streak})
    if not all_streaks: return pd.DataFrame()
    df = pd.DataFrame(all_streaks)
    return df.groupby(['direction','length']).agg(count=('year','count')).reset_index()

def return_dist(data, years, series='out', period='P1'):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        p = sub[(sub['period']==period) & sub['px_pct'].notna()].copy()
        p['year'] = yr
        rows.append(p[['year','bd','px_pct','px_chg','close_price','Volume','Open Interest','daily_range']])
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()

def autocorr_stats(data, years, series='out', period='P1', lags=5):
    results = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        p = sub[(sub['period']==period) & sub['px_pct'].notna()].sort_values('date')
        if len(p) < 10: continue
        for lag in range(1, lags+1):
            ac = p['px_pct'].autocorr(lag=lag)
            results.append({'year': yr, 'lag': lag, 'autocorr': ac})
    if not results: return pd.DataFrame()
    df = pd.DataFrame(results)
    return df.groupby('lag')['autocorr'].agg(['mean','std']).reset_index()

def vol_anomaly_stats(data, years, series='out', period='P1'):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        p = sub[(sub['period']==period) & sub['Volume'].notna() & sub['px_chg'].notna()].copy()
        if len(p) < 3: continue
        avg_vol = p['Volume'].mean()
        p['vol_z'] = (p['Volume'] - avg_vol) / (p['Volume'].std() + 1e-9)
        p['is_anomaly'] = p['vol_z'] > 1.5
        p['year'] = yr
        rows.append(p[['year','bd','date','Volume','vol_z','is_anomaly','px_chg','px_pct','close_price']])
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()

def cross_corr(data, years, period='P1'):
    rows = []
    for yr in [y for y in years if y in data]:
        out = data[yr]['out']; hk = data[yr]['hk']; fly = data[yr].get('fly', pd.DataFrame())
        for df, lbl in [(out,'out'),(hk,'hk'),(fly,'fly')]:
            if len(df) == 0: continue
            p = df[(df['period']==period) & df['close_price'].notna()][['date','close_price','Volume']].copy()
            p.columns = ['date', f'{lbl}_px', f'{lbl}_vol']
            rows.append((yr, lbl, p))
    by_year = {}
    for yr, lbl, p in rows:
        if yr not in by_year: by_year[yr] = pd.DataFrame({'date': []})
        if len(by_year[yr]) == 0:
            by_year[yr] = p
        else:
            by_year[yr] = pd.merge(by_year[yr], p, on='date', how='outer')
    merged_all = pd.concat([v for v in by_year.values() if len(v) > 0], ignore_index=True)
    px_cols = [c for c in merged_all.columns if c.endswith('_px')]
    if len(px_cols) < 2: return pd.DataFrame()
    return merged_all[px_cols].corr()

def week_stats(data, years, series='out', period='P1'):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        p = sub[(sub['period']==period) & sub['close_price'].notna()].copy()
        p['year'] = yr
        rows.append(p)
    if not rows: return pd.DataFrame()
    combined = pd.concat(rows, ignore_index=True)
    return combined.groupby(['week_in_period','year']).agg(
        week_open=('close_price','first'), week_close=('close_price','last'),
        avg_vol=('Volume','mean'), avg_range=('daily_range','mean'),
        avg_oi=('Open Interest','mean'), n_days=('bd','count')
    ).reset_index().assign(
        week_move=lambda x: x['week_close'] - x['week_open'],
        week_pct=lambda x: (x['week_close'] - x['week_open']) / x['week_open'].abs() * 100
    )

def spread_move_vs_roll(data, years, series='hk'):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get(series, pd.DataFrame())
        if len(sub) == 0: continue
        p2 = sub[(sub['period']=='P2') & sub['close_price'].notna()].sort_values('date').copy()
        p2['year'] = yr
        rows.append(p2)
    if not rows: return pd.DataFrame()
    combined = pd.concat(rows, ignore_index=True)
    return combined.groupby('bd').agg(
        avg_chg=('px_chg','mean'), std_chg=('px_chg','std'),
        n=('year','count'), up=('px_chg', lambda x: (x>0).sum())
    ).reset_index()

def bd_to_date_map(data, years):
    rows = []
    for yr in [y for y in years if y in data]:
        sub = data[yr].get('out', pd.DataFrame())
        if len(sub) == 0: continue
        b = data[yr]['bounds']
        for p in ['P1','P2','P3']:
            pp = sub[sub['period']==p].sort_values('date')
            for _, row in pp.iterrows():
                d = row['date']
                rows.append({
                    'Year': yr, 'Period': p,
                    'BD': int(row['bd']) if pd.notna(row['bd']) else None,
                    'Week': int(row['week_in_period']) if pd.notna(row.get('week_in_period')) else None,
                    'Date': str(d),
                    'Day': pd.Timestamp(d).strftime('%a'),
                    'Calendar Week': pd.Timestamp(d).isocalendar().week,
                })
    return pd.DataFrame(rows)

def current_bd_info(data, years):
    today = date_type.today()
    info = {'today': today, 'positions': []}
    for yr in sorted(years, reverse=True):
        b = data[yr]['bounds']
        sub = data[yr]['out']
        if len(sub) == 0: continue
        sub_d = sub[sub['close_price'].notna()].sort_values('date')
        # Find last row on or before today in this year's data
        past = sub_d[sub_d['date'].apply(lambda d: d <= today)]
        if len(past):
            last = past.iloc[-1]
            info['positions'].append({
                'year': yr, 'date': last['date'],
                'period': last['period'], 'bd': int(last['bd']) if pd.notna(last['bd']) else '?',
                'week': int(last['week_in_period']) if pd.notna(last.get('week_in_period')) else '?',
            })
    return info
